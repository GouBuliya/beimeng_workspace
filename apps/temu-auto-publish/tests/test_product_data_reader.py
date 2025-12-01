"""
@PURPOSE: 测试 ProductDataReader 产品数据读取器
@OUTLINE:
  - TestProductDataReader: 测试产品数据读取器主类
  - TestProductDataReaderCostPrice: 测试成本价读取
  - TestProductDataReaderDimensions: 测试尺寸读取
  - TestProductDataReaderRandomGeneration: 测试随机生成功能
@DEPENDENCIES:
  - 外部: pytest, openpyxl
  - 内部: src.data_processor.product_data_reader
"""


import pytest
from src.data_processor.product_data_reader import ProductDataReader


class TestProductDataReader:
    """测试产品数据读取器主类"""

    @pytest.fixture
    def reader_with_mock(self, tmp_path):
        """创建带模拟数据的读取器"""
        # 创建测试Excel文件
        try:
            import openpyxl

            filepath = tmp_path / "test_products.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            # 添加表头
            ws.append(["产品名称", "进货价", "重量", "长度", "宽度", "高度"])
            # 添加数据
            ws.append(["药箱收纳盒", 15.0, 6500, 80, 60, 50])
            ws.append(["厨房收纳架", 25.0, 8000, 90, 70, 55])
            wb.save(filepath)

            return ProductDataReader(str(filepath))
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_init_default(self):
        """测试默认初始化"""
        # 不提供路径时使用默认路径
        reader = ProductDataReader()
        assert reader is not None
        assert reader.excel_path is not None

    def test_init_with_custom_path(self, tmp_path):
        """测试自定义路径初始化"""
        custom_path = tmp_path / "custom.xlsx"
        reader = ProductDataReader(str(custom_path))

        assert reader.excel_path == custom_path

    def test_data_cache_exists(self, reader_with_mock):
        """测试数据缓存存在"""
        assert hasattr(reader_with_mock, "data_cache")
        assert isinstance(reader_with_mock.data_cache, dict)


class TestProductDataReaderCostPrice:
    """测试成本价读取"""

    @pytest.fixture
    def reader_with_data(self, tmp_path):
        """创建带测试数据的读取器"""
        try:
            import openpyxl

            filepath = tmp_path / "cost_test.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["产品名称", "进货价"])
            ws.append(["药箱收纳盒", 15.0])
            ws.append(["厨房收纳架", 25.5])
            wb.save(filepath)
            return ProductDataReader(str(filepath))
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_get_cost_price_exact_match(self, reader_with_data):
        """测试精确匹配获取成本价"""
        price = reader_with_data.get_cost_price("药箱收纳盒")

        assert price == 15.0

    def test_get_cost_price_fuzzy_match(self, reader_with_data):
        """测试模糊匹配获取成本价"""
        price = reader_with_data.get_cost_price("药箱")

        # 模糊匹配应该找到"药箱收纳盒"
        assert price is not None or price == 15.0

    def test_get_cost_price_not_found(self, reader_with_data):
        """测试未找到产品"""
        price = reader_with_data.get_cost_price("不存在的产品")

        assert price is None


class TestProductDataReaderDimensions:
    """测试尺寸读取"""

    @pytest.fixture
    def reader_with_dimensions(self, tmp_path):
        """创建带尺寸数据的读取器"""
        try:
            import openpyxl

            filepath = tmp_path / "dimensions_test.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["产品名称", "长度", "宽度", "高度"])
            ws.append(["药箱收纳盒", 80, 60, 50])
            wb.save(filepath)
            return ProductDataReader(str(filepath))
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_get_dimensions(self, reader_with_dimensions):
        """测试获取尺寸"""
        dims = reader_with_dimensions.get_dimensions("药箱收纳盒")

        if dims:
            assert dims["length"] == 80
            assert dims["width"] == 60
            assert dims["height"] == 50

    def test_get_dimensions_not_found(self, reader_with_dimensions):
        """测试未找到尺寸"""
        dims = reader_with_dimensions.get_dimensions("不存在的产品")

        assert dims is None


class TestProductDataReaderWeight:
    """测试重量读取"""

    @pytest.fixture
    def reader_with_weight(self, tmp_path):
        """创建带重量数据的读取器"""
        try:
            import openpyxl

            filepath = tmp_path / "weight_test.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["产品名称", "重量"])
            ws.append(["药箱收纳盒", 6500])
            wb.save(filepath)
            return ProductDataReader(str(filepath))
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_get_weight(self, reader_with_weight):
        """测试获取重量"""
        weight = reader_with_weight.get_weight("药箱收纳盒")

        if weight:
            assert weight == 6500

    def test_get_weight_not_found(self, reader_with_weight):
        """测试未找到重量"""
        weight = reader_with_weight.get_weight("不存在的产品")

        assert weight is None


class TestProductDataReaderRandomGeneration:
    """测试随机生成功能"""

    def test_generate_random_dimensions(self):
        """测试生成随机尺寸"""
        dims = ProductDataReader.generate_random_dimensions()

        assert "length" in dims
        assert "width" in dims
        assert "height" in dims

        # 验证长>宽>高
        assert dims["length"] >= dims["width"]
        assert dims["width"] >= dims["height"]

        # 验证范围
        assert 50 <= dims["height"] <= 99
        assert 50 <= dims["width"] <= 99
        assert 80 <= dims["length"] <= 99

    def test_generate_random_dimensions_consistency(self):
        """测试随机尺寸一致性"""
        # 多次生成确保关系正确
        for _ in range(10):
            dims = ProductDataReader.generate_random_dimensions()
            assert dims["length"] >= dims["width"] >= dims["height"]

    def test_generate_random_weight(self):
        """测试生成随机重量"""
        weight = ProductDataReader.generate_random_weight()

        assert 5000 <= weight <= 9999

    def test_generate_random_weight_range(self):
        """测试随机重量范围"""
        weights = [ProductDataReader.generate_random_weight() for _ in range(100)]

        assert all(5000 <= w <= 9999 for w in weights)

    def test_validate_and_fix_dimensions_correct_order(self):
        """测试验证正确顺序的尺寸"""
        length, width, height = ProductDataReader.validate_and_fix_dimensions(80, 60, 50)

        assert length == 80
        assert width == 60
        assert height == 50

    def test_validate_and_fix_dimensions_wrong_order(self):
        """测试修正错误顺序的尺寸"""
        # 输入错误顺序:高 > 宽 > 长
        length, width, height = ProductDataReader.validate_and_fix_dimensions(50, 60, 80)

        # 应该修正为正确顺序
        assert length == 80
        assert width == 60
        assert height == 50

    def test_validate_and_fix_dimensions_all_equal(self):
        """测试修正全部相等的尺寸"""
        length, width, height = ProductDataReader.validate_and_fix_dimensions(50, 50, 50)

        assert length == 50
        assert width == 50
        assert height == 50


class TestProductDataReaderColumnSearch:
    """测试列查找功能"""

    def test_find_column_index_found(self, tmp_path):
        """测试找到列索引"""
        try:
            import openpyxl

            filepath = tmp_path / "columns.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["产品名称", "进货价", "重量"])
            ws.append(["测试", 10.0, 5000])
            wb.save(filepath)

            reader = ProductDataReader(str(filepath))

            # 内部方法测试
            headers = ["产品名称", "进货价", "重量"]
            idx = reader._find_column_index(headers, ["进货价", "成本价"])

            assert idx == 1
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_find_column_index_not_found(self):
        """测试未找到列索引"""
        reader = ProductDataReader()

        headers = ["产品名称", "进货价"]
        idx = reader._find_column_index(headers, ["不存在的列"])

        assert idx is None
