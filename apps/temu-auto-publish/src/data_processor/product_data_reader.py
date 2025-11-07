"""
@PURPOSE: 从Excel文件（10月品.xlsx）读取产品数据，用于批量编辑
@OUTLINE:
  - class ProductDataReader: 产品数据读取器
  - get_cost_price(): 读取产品成本价
  - get_dimensions(): 读取产品尺寸信息
  - get_weight(): 读取产品重量信息
  - generate_random_dimensions(): 生成随机尺寸（50-99cm，长>宽>高）
  - generate_random_weight(): 生成随机重量（5000-9999G）
@TECH_DEBT:
  - TODO: 添加缓存机制避免重复读取Excel
  - TODO: 支持多个Excel文件
@DEPENDENCIES:
  - 外部: openpyxl, pandas (可选)
"""
import random
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from packages.common.logger import logger

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel数据读取功能不可用")


class ProductDataReader:
    """产品数据读取器，从 `data/input/10月品 .xlsx` 读取产品信息."""

    def __init__(self, excel_path: Optional[str] = None) -> None:
        """初始化数据读取器.

        Args:
            excel_path: Excel文件路径，默认定位到应用目录下的 `data/input/10月品 .xlsx`.
        """
        if excel_path:
            self.excel_path = Path(excel_path)
        else:
            app_root = Path(__file__).resolve().parents[2]
            self.excel_path = app_root / "data/input/10月品 .xlsx"

        self.data_cache = {}
        self._load_data()
    
    def _load_data(self):
        """加载Excel数据到缓存."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl 未安装，无法读取Excel数据")
            return
        
        if not self.excel_path.exists():
            logger.warning(f"Excel文件不存在: {self.excel_path}")
            return
        
        try:
            logger.info(f"正在加载Excel数据: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            logger.debug(f"找到 {len(headers)} 列数据")
            
            # 找到关键列的索引
            self.col_product_name = self._find_column_index(headers, ['产品名称', '商品名称', '名称'])
            self.col_cost_price = self._find_column_index(headers, ['进货价', '成本价', '采购价'])
            self.col_weight = self._find_column_index(headers, ['重量', '毛重', 'weight'])
            self.col_length = self._find_column_index(headers, ['长度', '长', 'length'])
            self.col_width = self._find_column_index(headers, ['宽度', '宽', 'width'])
            self.col_height = self._find_column_index(headers, ['高度', '高', 'height'])
            
            # 读取所有产品数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and self.col_product_name is not None:
                    product_name = row[self.col_product_name]
                    if product_name and isinstance(product_name, str):
                        self.data_cache[product_name.strip()] = {
                            'cost_price': self._get_cell_value(row, self.col_cost_price),
                            'weight': self._get_cell_value(row, self.col_weight),
                            'length': self._get_cell_value(row, self.col_length),
                            'width': self._get_cell_value(row, self.col_width),
                            'height': self._get_cell_value(row, self.col_height),
                        }
            
            wb.close()
            logger.success(f"✓ 已加载 {len(self.data_cache)} 个产品数据")
            
        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
    
    def _find_column_index(self, headers: list, keywords: list) -> Optional[int]:
        """根据关键词查找列索引.
        
        Args:
            headers: 表头列表
            keywords: 可能的列名关键词
            
        Returns:
            列索引，未找到返回None
        """
        for keyword in keywords:
            for i, header in enumerate(headers):
                if header and keyword in str(header):
                    return i
        return None
    
    def _get_cell_value(self, row: tuple, col_index: Optional[int]) -> Optional[float]:
        """获取单元格值并转换为数字.
        
        Args:
            row: 行数据
            col_index: 列索引
            
        Returns:
            数字值，无效返回None
        """
        if col_index is None or col_index >= len(row):
            return None
        
        value = row[col_index]
        if value is None:
            return None
        
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    
    def get_cost_price(self, product_name: str) -> Optional[float]:
        """获取产品成本价.
        
        Args:
            product_name: 产品名称
            
        Returns:
            成本价（元），未找到返回None
        """
        if not self.data_cache:
            return None
        
        # 精确匹配
        if product_name in self.data_cache:
            return self.data_cache[product_name].get('cost_price')
        
        # 模糊匹配（产品名包含关键词）
        for name, data in self.data_cache.items():
            if product_name in name or name in product_name:
                cost_price = data.get('cost_price')
                if cost_price:
                    logger.debug(f"模糊匹配到产品: {name}, 成本价: {cost_price}")
                    return cost_price
        
        return None
    
    def get_dimensions(self, product_name: str) -> Optional[Dict[str, int]]:
        """获取产品尺寸信息.
        
        Args:
            product_name: 产品名称
            
        Returns:
            尺寸字典 {'length': int, 'width': int, 'height': int}，未找到返回None
        """
        if not self.data_cache:
            return None
        
        # 精确匹配
        data = self.data_cache.get(product_name)
        if not data:
            # 模糊匹配
            for name, d in self.data_cache.items():
                if product_name in name or name in product_name:
                    data = d
                    break
        
        if data:
            length = data.get('length')
            width = data.get('width')
            height = data.get('height')
            
            if length and width and height:
                return {
                    'length': int(length),
                    'width': int(width),
                    'height': int(height)
                }
        
        return None
    
    def get_weight(self, product_name: str) -> Optional[int]:
        """获取产品重量.
        
        Args:
            product_name: 产品名称
            
        Returns:
            重量（克），未找到返回None
        """
        if not self.data_cache:
            return None
        
        # 精确匹配
        data = self.data_cache.get(product_name)
        if not data:
            # 模糊匹配
            for name, d in self.data_cache.items():
                if product_name in name or name in product_name:
                    data = d
                    break
        
        if data:
            weight = data.get('weight')
            if weight:
                return int(weight)
        
        return None
    
    @staticmethod
    def generate_random_dimensions() -> Dict[str, int]:
        """生成随机尺寸（50-99cm，确保长>宽>高）.
        
        Returns:
            尺寸字典 {'length': int, 'width': int, 'height': int}
        """
        # 生成长度（80-99cm）
        length = random.randint(80, 99)
        
        # 生成宽度（比长度小10-20cm）
        width = random.randint(max(50, length - 20), length - 5)
        
        # 生成高度（比宽度小10-20cm）
        height = random.randint(max(50, width - 20), width - 5)
        
        logger.debug(f"生成随机尺寸: 长={length}cm, 宽={width}cm, 高={height}cm")
        
        return {
            'length': length,
            'width': width,
            'height': height
        }
    
    @staticmethod
    def generate_random_weight() -> int:
        """生成随机重量（5000-9999G）.
        
        Returns:
            重量（克）
        """
        weight = random.randint(5000, 9999)
        logger.debug(f"生成随机重量: {weight}G")
        return weight
    
    @staticmethod
    def validate_and_fix_dimensions(length: int, width: int, height: int) -> Tuple[int, int, int]:
        """验证并修正尺寸，确保长>宽>高.
        
        Args:
            length: 长度
            width: 宽度
            height: 高度
            
        Returns:
            修正后的尺寸元组 (length, width, height)
        """
        # 将三个值排序，最大的是长，最小的是高
        dimensions = sorted([length, width, height], reverse=True)
        
        fixed_length, fixed_width, fixed_height = dimensions
        
        if (fixed_length, fixed_width, fixed_height) != (length, width, height):
            logger.warning(
                f"尺寸不符合长>宽>高规则，已自动调整: "
                f"({length}, {width}, {height}) -> ({fixed_length}, {fixed_width}, {fixed_height})"
            )
        
        return fixed_length, fixed_width, fixed_height

